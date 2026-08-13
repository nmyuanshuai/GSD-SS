# ==================== 内蒙古80年代草普录入 - 最终完整修复版 ====================
# 修复内容：
# 1. 表7小计求和
# 2. 表67录入：从M列为空的行开始写入（使用openpyxl读取公式值）
# 3. 表345录入：跳过已有灌木植物行，从空行开始写草本
# 4. 表4干重提取：
#    a. 保留原有规则，但改用openpyxl data_only=True读取公式计算值
#    b. 如果原有规则提取不到干重，则遍历找到所有干重列，取左数第二个干重列
# 5. 表1录入：新规则（扫描分散单元格，D盟市/E旗县/F乡镇）
# 6. 删除空行（快速版）
# 
# 新增修复（2026-05-08）：
# 7. 表1文件名匹配时自动去除首尾空格（解决文件名空格问题）
# 8. W列干重保留2位小数，重量合计也保留2位小数
# 9. 表67录入时，样地行的Y、Z、AA列求和（AB列不求）
# 10. 表3特殊指标各自独立提取，遇到其他指标关键词停止（解决张冠李戴）
# 11. 表5频度数据确保写入所有植物（修复new_plants5写入逻辑）
#
# 新增修复（2026-05-09）：
# 12. 解决合并单元格写入报错（添加safe_write_cell函数）
# 13. 保护前9行表头，从第10行开始操作（防止覆盖标题）

import os
import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
import warnings
warnings.filterwarnings('ignore')

# ==================== 配置路径 ====================
WORK_DIR = r"F:\内蒙古80年代草普录入\呼伦贝尔市候补"
TARGET_EXCEL = r"F:\内蒙古80年代草普录入\各个盟市用代码录入\呼伦贝尔市80草普录入.xlsx"
LOG_FILE = r"F:\内蒙古80年代草普录入\各个盟市用代码录入\处理日志.txt"
HEADER_ROWS = 3  # 前3行为表头，保护不操作
# ================================================

print("=" * 70)
print("内蒙古80年代草普录入数据处理系统 - 最终完整修复版")
print(f"保护前 {HEADER_ROWS} 行表头，从第 {HEADER_ROWS + 1} 行开始操作")
print("顺序：表7求和 -> 表67录入 -> 表345录入 -> 表1录入 -> 删除空行")
print("表1字段：样地号(C) | 盟市(D) | 旗县(E) | 乡镇(F) | 经度(G) | 纬度(H) | 海拔(I) | 日期(J) | 优势度定名(K)")
print("=" * 70)
print(f"工作目录: {WORK_DIR}")
print(f"目标文件: {TARGET_EXCEL}")
print("=" * 70)

# ==================== 合并单元格安全写入函数 ====================
def is_merged_cell(ws, row, col):
    """检查单元格是否为合并单元格的非左上角"""
    for merged_range in ws.merged_cells:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return not (row == merged_range.min_row and col == merged_range.min_col)
    return False

def safe_write_cell(ws, row, col, value):
    """安全写入单元格，自动处理合并单元格"""
    try:
        ws.cell(row=row, column=col, value=value)
    except AttributeError:
        for merged_range in ws.merged_cells:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                return

# ==================== 智能选择sheet函数 ====================
def select_best_sheet(wb, target_names):
    all_sheets = wb.sheetnames
    
    for sheet in all_sheets:
        for name in target_names:
            if name in sheet and ('汇总' in sheet or '汇总' in name):
                print(f"    选择包含'汇总'的sheet: {sheet}")
                return sheet
    
    for sheet in all_sheets:
        for name in target_names:
            if sheet == name or sheet == name + '汇总':
                print(f"    选择精确匹配的sheet: {sheet}")
                return sheet
    
    for sheet in all_sheets:
        for name in target_names:
            if name in sheet:
                clean_sheet = re.sub(r'[（(][^）)]*[）)]', '', sheet).strip()
                if clean_sheet == name or clean_sheet == name + '汇总':
                    print(f"    选择基础名称的sheet（去除括号后匹配）: {sheet}")
                return sheet
    
    for sheet in all_sheets:
        for name in target_names:
            if name in sheet or sheet in name:
                print(f"    选择匹配的sheet: {sheet}")
                return sheet
    
    if len(all_sheets) == 1:
        print(f"    只有唯一sheet: {all_sheets[0]}")
        return all_sheets[0]
    
    return None

def read_excel_sheet_smart(file_path, target_names):
    """智能读取Excel的sheet（使用pandas）"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_name = select_best_sheet(wb, target_names)
        if sheet_name is None:
            wb.close()
            return None
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        wb.close()
        return df
    except Exception as e:
        return None

def read_excel_sheet_openpyxl(file_path, target_names):
    """用openpyxl读取Excel，data_only=True获取公式计算值"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_name = select_best_sheet(wb, target_names)
        if sheet_name is None:
            wb.close()
            return None
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([cell if cell is not None else '' for cell in row])
        wb.close()
        return pd.DataFrame(data)
    except Exception as e:
        return None

# ==================== 日志函数 ====================
def write_log(message, print_console=True):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_msg = f"[{timestamp}] {message}"
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_msg + "\n")
    if print_console:
        print(log_msg)

# ==================== 去掉.0的通用函数 ====================
def remove_decimal_zero(val):
    if val is None or pd.isna(val):
        return val
    if isinstance(val, float):
        if val.is_integer():
            return int(val)
        return val
    if isinstance(val, str):
        if val.endswith('.0') and val[:-2].replace('.', '').isdigit():
            return val[:-2]
        val = re.sub(r'(\d+)\.0([°′″])', r'\1\2', val)
        return val
    return val

# ==================== 表1新提取函数 ====================
def scan_right(df, row_idx, col_idx, max_offset=20, stop_keywords=None):
    if stop_keywords is None:
        stop_keywords = []
    for offset in range(1, max_offset + 1):
        check_col = col_idx + offset
        if check_col >= df.shape[1]:
            break
        val = df.iloc[row_idx, check_col]
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        if val_str == '' or val_str == 'nan':
            continue
        for kw in stop_keywords:
            if kw in val_str:
                return None
        return val_str
    return None

def scan_left(df, row_idx, col_idx, max_offset=10, stop_keywords=None):
    if stop_keywords is None:
        stop_keywords = []
    for offset in range(1, max_offset + 1):
        check_col = col_idx - offset
        if check_col < 0:
            break
        val = df.iloc[row_idx, check_col]
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        if val_str == '' or val_str == 'nan':
            continue
        for kw in stop_keywords:
            if kw in val_str:
                return None
        return val_str
    return None

def extract_longitude_latitude(df, row_idx, col_idx):
    numbers = []
    has_deg = False
    has_min = False
    has_sec = False
    stop_words = ['海拔高度', '坡度', '记载面积', '调查人', '行政地区', '样地号', '日期']
    for offset in range(1, 30):
        check_col = col_idx + offset
        if check_col >= df.shape[1]:
            break
        val = df.iloc[row_idx, check_col]
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        if any(stop in val_str for stop in stop_words):
            break
        if val_str in ['°', '【°】']:
            has_deg = True
        elif val_str in ['′', '【′】']:
            has_min = True
        elif val_str in ['″', '【″】']:
            has_sec = True
        else:
            num_match = re.search(r'(\d+\.?\d*)', val_str)
            if num_match:
                numbers.append(num_match.group(1))
    parts = []
    if len(numbers) > 0:
        parts.append(str(int(float(numbers[0]))) + "°")
    elif has_deg:
        parts.append("°")
    elif has_min or has_sec:
        parts.append("°")
    if len(numbers) > 1:
        parts.append(str(int(float(numbers[1]))) + "′")
    elif has_min:
        parts.append("′")
    elif has_sec and not parts:
        if '°' not in ''.join(parts):
            parts.append("°")
        parts.append("′")
    if len(numbers) > 2:
        parts.append(str(int(float(numbers[2]))) + "″")
    elif has_sec:
        parts.append("″")
    return ''.join(parts) if parts else None

def extract_date(df, row_idx, col_idx):
    year_val = None
    month_val = None
    day_val = None
    has_year = False
    has_month = False
    has_day = False
    stop_words = ['海拔高度', '记载面积', '调查人', '样地号']
    for offset in range(1, 20):
        check_col = col_idx + offset
        if check_col >= df.shape[1]:
            break
        val = df.iloc[row_idx, check_col]
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        if any(stop in val_str for stop in stop_words):
            break
        if val_str in ['年', '【年】']:
            has_year = True
        elif val_str in ['月', '【月】']:
            has_month = True
        elif val_str in ['日', '【日】']:
            has_day = True
        else:
            num_match = re.search(r'(\d+\.?\d*)', val_str)
            if num_match:
                num = num_match.group(1)
                if year_val is None and len(num) == 4:
                    year_val = num
                elif month_val is None and 1 <= float(num) <= 12:
                    month_val = num
                elif day_val is None and 1 <= float(num) <= 31:
                    day_val = num
    parts = []
    if year_val:
        parts.append(str(int(float(year_val))) + "年")
    elif has_year:
        parts.append("年")
    if month_val:
        parts.append(str(int(float(month_val))).zfill(2) + "月")
    elif has_month:
        parts.append("月")
    if day_val:
        parts.append(str(int(float(day_val))).zfill(2) + "日")
    elif has_day:
        parts.append("日")
    return ''.join(parts) if parts else None

def extract_table1_info(file_path):
    info = {'样地号': None, '经度': None, '纬度': None, '日期': None, 
            '盟市': None, '旗县': None, '乡镇': None, '海拔': None, '优势度定名': None}
    FIELD_STOP_WORDS = ['样地号：', '日期：', '经度：', '纬度：', '盟（市）', '盟(市)', 
                       '旗（县）', '旗(县)', '公社：', '乡镇：', '海拔高度：', '优势度定名', 
                       '野外定名', '行政地区：']
    try:
        possible_sheets = ['表1', '表一', 'Sheet1', '表', 'Sheet', '数据']
        df = None
        for sheet_name in possible_sheets:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                break
            except:
                continue
        if df is None:
            return info
        for row_idx in range(min(30, len(df))):
            for col_idx in range(min(40, df.shape[1])):
                cell = df.iloc[row_idx, col_idx]
                if pd.isna(cell):
                    continue
                cell_str = str(cell).strip()
                if '样地号：' in cell_str and info['样地号'] is None:
                    value = scan_right(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        num_match = re.search(r'(\d+\.?\d*)', value)
                        if num_match:
                            info['样地号'] = remove_decimal_zero(num_match.group(1))
                if ('盟（市）' in cell_str or '盟(市)' in cell_str) and info['盟市'] is None:
                    value = scan_left(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        info['盟市'] = value
                if ('旗（县）' in cell_str or '旗(县)' in cell_str) and info['旗县'] is None:
                    value = scan_left(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        info['旗县'] = value
                if ('公社：' in cell_str or '乡镇：' in cell_str) and info['乡镇'] is None:
                    value = scan_left(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        info['乡镇'] = value
                if ('海拔高度：' in cell_str or '海拔：' in cell_str) and info['海拔'] is None:
                    value = scan_right(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        num_match = re.search(r'(\d+\.?\d*)', value)
                        if num_match:
                            info['海拔'] = remove_decimal_zero(num_match.group(1))
                if ('优势度定名' in cell_str or '野外定名' in cell_str) and info['优势度定名'] is None:
                    value = scan_right(df, row_idx, col_idx, 10, FIELD_STOP_WORDS)
                    if value:
                        info['优势度定名'] = value
                if ('经度：' in cell_str or '东经：' in cell_str) and info['经度'] is None:
                    result = extract_longitude_latitude(df, row_idx, col_idx)
                    if result is not None:
                        info['经度'] = result
                if ('纬度：' in cell_str or '北纬：' in cell_str) and info['纬度'] is None:
                    result = extract_longitude_latitude(df, row_idx, col_idx)
                    if result is not None:
                        info['纬度'] = result
                if '日期：' in cell_str and info['日期'] is None:
                    result = extract_date(df, row_idx, col_idx)
                    if result:
                        info['日期'] = result
        return info
    except Exception as e:
        return info

# ==================== 通用函数 ====================
def extract_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = re.sub(r'[^\d\.\-]', '', val.strip())
        try:
            return float(s) if '.' in s else int(s)
        except:
            return None
    return None

def clean_text(text):
    if pd.isna(text):
        return ""
    return re.sub(r'\s+', '', str(text)).lower()

def find_column_by_keyword(df, keyword, start_row=0, end_row=50):
    clean_keyword = clean_text(keyword)
    for row_idx in range(start_row, min(end_row, len(df))):
        row = df.iloc[row_idx]
        for col_idx, cell in enumerate(row):
            if pd.notna(cell) and clean_keyword in clean_text(cell):
                return col_idx, row_idx
    return None, None

def find_columns_by_keyword(df, keyword, start_row=0, end_row=50):
    columns = []
    clean_keyword = clean_text(keyword)
    for row_idx in range(start_row, min(end_row, len(df))):
        row = df.iloc[row_idx]
        for col_idx, cell in enumerate(row):
            if pd.notna(cell) and clean_keyword in clean_text(cell):
                columns.append((col_idx, row_idx))
    return columns

def extract_number_right(df, row_idx, col_idx, max_offset=10):
    for offset in range(1, max_offset + 1):
        check_col = col_idx + offset
        if check_col < len(df.columns):
            cell = df.iloc[row_idx, check_col]
            if pd.notna(cell):
                try:
                    if isinstance(cell, (int, float)):
                        return float(cell)
                except:
                    pass
                num = re.sub(r'[^\d\.\-]', '', str(cell))
                if num and num != "-":
                    try:
                        return float(num)
                    except:
                        continue
    return None

def extract_number_right_for_metrics(df, row_idx, col_idx, current_keyword, max_offset=15):
    stop_keywords = ["自然生殖枝高", "自然叶层高", "总盖度", "基盖度"]
    for offset in range(1, max_offset + 1):
        check_col = col_idx + offset
        if check_col >= df.shape[1]:
            break
        cell = df.iloc[row_idx, check_col]
        if pd.isna(cell):
            continue
        cell_str = str(cell).strip()
        for kw in stop_keywords:
            if kw in cell_str and kw != current_keyword:
                return None
        num_match = re.search(r'(\d+\.?\d*)', cell_str)
        if num_match:
            return float(num_match.group(1))
    return None

def is_valid_plant_name(name):
    if not name or not isinstance(name, str):
        return False
    name = name.strip()
    if not name:
        return False
    name_no_space = re.sub(r'\s+', '', name)
    if '合计' in name_no_space or '小计' in name_no_space:
        return False
    if re.fullmatch(r'[\d\.]+', name_no_space):
        return False
    if not any('\u4e00' <= c <= '\u9fff' for c in name):
        return False
    exclude = ['样地', '植物名称', '中名', '植   物   名  称', '盖度', '高度', '多度', '频度']
    return name not in exclude

def get_plant_data_mapping(df, name_col, name_row, data_col, max_rows=100):
    mapping = {}
    if data_col is None or name_col is None:
        return mapping
    row_idx = name_row + 1
    end_row = min(name_row + max_rows, len(df))
    while row_idx < end_row:
        if name_col >= len(df.columns):
            break
        plant = df.iloc[row_idx, name_col]
        if pd.notna(plant):
            plant_str = str(plant).strip()
            if is_valid_plant_name(plant_str):
                if data_col < len(df.columns):
                    val = df.iloc[row_idx, data_col]
                    if pd.notna(val):
                        try:
                            mapping[plant_str] = float(val)
                        except:
                            mapping[plant_str] = val
        row_idx += 1
    return mapping

def get_all_plant_names_until_next_title(df, name_col, start_row, title_keywords, max_rows=200):
    plants = []
    row_idx = start_row
    end_row = min(start_row + max_rows, len(df))
    while row_idx < end_row:
        if name_col >= len(df.columns):
            break
        is_next_title = False
        for col in range(min(20, len(df.columns))):
            cell = str(df.iloc[row_idx, col]) if pd.notna(df.iloc[row_idx, col]) else ""
            for kw in title_keywords:
                if kw in cell and len(cell) > 5:
                    is_next_title = True
                    break
            if is_next_title:
                break
        if is_next_title:
            break
        cell = df.iloc[row_idx, name_col]
        cell_str = str(cell).strip() if pd.notna(cell) else ""
        if '小计' in cell_str or '合计' in cell_str or '总计' in cell_str:
            break
        if cell_str == "":
            empty_count = 0
            for c in range(max(0, name_col-2), min(len(df.columns), name_col+3)):
                if pd.isna(df.iloc[row_idx, c]) or str(df.iloc[row_idx, c]).strip() == "":
                    empty_count += 1
            if empty_count >= 3:
                break
            row_idx += 1
            continue
        if is_valid_plant_name(cell_str):
            plants.append(cell_str)
        row_idx += 1
    return plants

def find_cover_column(df, start_row=0, end_row=50):
    for row_idx in range(start_row, min(end_row, len(df))):
        for col_idx in range(min(25, len(df.columns))):
            cell = df.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell).strip()
                cell_clean = re.sub(r'\s+', '', cell_str)
                if '盖度' in cell_clean and '总' not in cell_clean and '基' not in cell_clean:
                    if cell_clean != '总盖度' and cell_clean != '基盖度':
                        return col_idx, row_idx
    return None, None

# ==================== 表3处理函数 ====================
def process_table3(source_file):
    df3 = read_excel_sheet_smart(source_file, ['表3', '表三', '3', '表3汇总', '表三汇总', '3汇总'])
    if df3 is None:
        return None, None, None
    title_row = None
    for row_idx in range(min(500, len(df3))):
        for col_idx in range(min(30, len(df3.columns))):
            cell = df3.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                if "草本半灌木层样方记载表" in cell_str:
                    title_row = row_idx
                    print(f"    表3: 找到标题在第{title_row+1}行")
                    break
        if title_row is not None:
            break
    if title_row is None:
        print(f"    表3: 未找到标题")
        return None, None, None
    end_row = len(df3)
    next_titles = ["草本半灌木层产量草测定表", "记名样方表"]
    for row_idx in range(title_row + 1, min(1000, len(df3))):
        for col_idx in range(min(30, len(df3.columns))):
            cell = df3.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                for nt in next_titles:
                    if nt in cell_str:
                        end_row = row_idx
                        break
                if end_row < len(df3):
                    break
        if end_row < len(df3):
            break
    name_col, name_row = find_column_by_keyword(df3, "中名", title_row, end_row)
    if name_col is None:
        name_col, name_row = find_column_by_keyword(df3, "植物名称", title_row, end_row)
    if name_col is None:
        return None, None, None
    title_keywords_for_stop = ["内蒙古自治区草场资源调查", "表", "产量草测定表", "记名样方表"]
    plant_names = get_all_plant_names_until_next_title(df3, name_col, name_row + 1, title_keywords_for_stop, 200)
    height_columns_info = find_columns_by_keyword(df3, "高度", title_row, end_row)
    height_columns = [col for col, row in sorted(height_columns_info, key=lambda x: x[0])]
    height_cols_4 = height_columns[:4]
    cover_col, cover_row = find_cover_column(df3, title_row, end_row)
    if cover_col is None:
        cover_col, cover_row = find_column_by_keyword(df3, "盖度", title_row, end_row)
    abund_col, abund_row = find_column_by_keyword(df3, "多度", title_row, end_row)
    if abund_col is None:
        abund_col, abund_row = find_column_by_keyword(df3, "密度", title_row, end_row)
    special_metrics = {}
    for keyword, target_col in [("自然生殖枝高", 17), ("自然叶层高", 19), ("总盖度", 21)]:
        col_idx, row_idx = find_column_by_keyword(df3, keyword, title_row, end_row)
        if col_idx is not None:
            val = extract_number_right_for_metrics(df3, row_idx, col_idx, keyword)
            if val is not None:
                special_metrics[keyword] = {"value": val, "target_col": target_col}
                print(f"    提取{keyword}: {val}")
    return df3, plant_names, {
        "name_col": name_col, "name_row": name_row,
        "height_cols": height_cols_4, "cover_col": cover_col,
        "abund_col": abund_col, "special_metrics": special_metrics,
        "title_row": title_row, "end_row": end_row
    }

# ==================== 表4处理函数 ====================
def process_table4(source_file):
    df4 = read_excel_sheet_openpyxl(source_file, ['表4', '表四', '4', '表4汇总', '表四汇总', '4汇总'])
    if df4 is None:
        return None, None, None
    
    title_row = None
    for row_idx in range(min(500, len(df4))):
        for col_idx in range(min(30, len(df4.columns))):
            cell = df4.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                if "产量草测定表" in cell_str or "表四" in cell_str:
                    title_row = row_idx
                    print(f"    表4: 找到标题在第{title_row+1}行")
                    break
        if title_row is not None:
            break
    
    if title_row is None:
        title_row = 0
    
    end_row = len(df4)
    next_titles = ["样方记载表", "记名样方表", "表3", "表5"]
    for row_idx in range(title_row + 1, min(1000, len(df4))):
        for col_idx in range(min(30, len(df4.columns))):
            cell = df4.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                for nt in next_titles:
                    if nt in cell_str:
                        end_row = row_idx
                        break
                if end_row < len(df4):
                    break
        if end_row < len(df4):
            break
    
    name_col, name_row = find_column_by_keyword(df4, "中名", title_row, end_row)
    if name_col is None:
        name_col, name_row = find_column_by_keyword(df4, "植物名称", title_row, end_row)
    if name_col is None:
        return None, None, None
    
    title_keywords_for_stop = ["内蒙古自治区草场资源调查", "表3", "表5", "样方记载表"]
    plant_names = get_all_plant_names_until_next_title(df4, name_col, name_row + 1, title_keywords_for_stop, 150)
    
    weight_col = None
    
    avg_col, avg_row = find_column_by_keyword(df4, "一平方米平均", title_row, end_row)
    if avg_col is None:
        avg_col, avg_row = find_column_by_keyword(df4, "克/㎡", title_row, end_row)
    
    if avg_col is not None and avg_row + 1 < len(df4):
        if avg_col < len(df4.columns):
            cell = df4.iloc[avg_row + 1, avg_col]
            if pd.notna(cell) and str(cell).strip() == '干重':
                weight_col = avg_col
                print(f"    表4: 原有规则找到干重列: 第{weight_col+1}列")
        if weight_col is None and avg_col + 1 < len(df4.columns):
            cell = df4.iloc[avg_row + 1, avg_col + 1]
            if pd.notna(cell) and str(cell).strip() == '干重':
                weight_col = avg_col + 1
                print(f"    表4: 原有规则找到干重列: 第{weight_col+1}列（右下方）")
    
    if weight_col is None:
        print(f"    表4: 原有规则未找到干重，启用回退方案...")
        dry_weight_columns = []
        for row_idx in range(title_row, min(title_row + 15, len(df4))):
            for col_idx in range(min(40, len(df4.columns))):
                cell = df4.iloc[row_idx, col_idx]
                if pd.notna(cell) and str(cell).strip() == '干重':
                    if col_idx not in dry_weight_columns:
                        dry_weight_columns.append(col_idx)
                        print(f"    找到干重列: 第{col_idx+1}列")
        
        if len(dry_weight_columns) >= 2:
            weight_col = dry_weight_columns[1]
            print(f"    选择左数第二个干重列: 第{weight_col+1}列")
        elif len(dry_weight_columns) == 1:
            weight_col = dry_weight_columns[0]
            print(f"    只有1个干重列，使用它: 第{weight_col+1}列")
    
    return df4, plant_names, {
        "name_col": name_col,
        "name_row": name_row,
        "weight_col": weight_col,
        "title_row": title_row,
        "end_row": end_row
    }

# ==================== 表5处理函数 ====================
def process_table5(source_file):
    df5 = read_excel_sheet_smart(source_file, ['表5', '表五', '5', '表5汇总', '表五汇总', '5汇总'])
    if df5 is None:
        return None, None, None
    title_row = None
    for row_idx in range(min(500, len(df5))):
        for col_idx in range(min(30, len(df5.columns))):
            cell = df5.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                if "记名样方表" in cell_str:
                    title_row = row_idx
                    break
        if title_row is not None:
            break
    if title_row is None:
        return None, None, None
    end_row = len(df5)
    next_titles = ["草本半灌木层样方记载表", "草本半灌木层产量草测定表"]
    for row_idx in range(title_row + 1, min(1000, len(df5))):
        for col_idx in range(min(30, len(df5.columns))):
            cell = df5.iloc[row_idx, col_idx]
            if pd.notna(cell):
                cell_str = str(cell)
                for nt in next_titles:
                    if nt in cell_str:
                        end_row = row_idx
                        break
                if end_row < len(df5):
                    break
        if end_row < len(df5):
            break
    name_col, name_row = find_column_by_keyword(df5, "中名", title_row, end_row)
    if name_col is None:
        name_col, name_row = find_column_by_keyword(df5, "植物名称", title_row, end_row)
    if name_col is None:
        return None, None, None
    title_keywords_for_stop = ["内蒙古自治区草场资源调查", "表", "样方记载表", "产量草测定表"]
    plant_names = get_all_plant_names_until_next_title(df5, name_col, name_row + 1, title_keywords_for_stop, 150)
    freq_col, freq_row = find_column_by_keyword(df5, "频度", title_row, end_row)
    return df5, plant_names, {
        "name_col": name_col, "name_row": name_row, 
        "freq_col": freq_col,
        "title_row": title_row, "end_row": end_row
    }

# ==================== 表6表7代码 ====================
TABLE_TITLE = "内蒙古自治区草场资源调查灌木及高大草本小样地记载表"

def get_cell_value_openpyxl(ws, row, col):
    if row < 1 or col < 1:
        return None
    try:
        for merged_range in ws.merged_cells:
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left.value
        return ws.cell(row=row, column=col).value
    except:
        return None

def safe_str_openpyxl(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).strip()
    s = re.sub(r'[\s\u3000\xa0\xa0]', '', s)
    return s

def extract_number_openpyxl(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r'(\d+\.?\d*)', str(value))
    if match:
        return float(match.group(1))
    return None

def find_first_section_openpyxl(ws):
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(10, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val and TABLE_TITLE in safe_str_openpyxl(val):
                return row
    return 1

def find_next_section_start_openpyxl(ws, current_start):
    for row in range(current_start + 1, min(ws.max_row + 1, 200)):
        for col in range(1, min(10, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val and TABLE_TITLE in safe_str_openpyxl(val):
                return row
    return ws.max_row + 1

def find_parent_child_column_openpyxl(ws, parent_keywords, child_keywords, start_row, end_row, max_col=50):
    for row in range(start_row, min(end_row, ws.max_row + 1)):
        for col in range(1, min(max_col, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val:
                val_str = safe_str_openpyxl(val)
                for pk in parent_keywords:
                    if pk in val_str:
                        for offset in range(1, 20):
                            if col + offset <= ws.max_column:
                                next_val = get_cell_value_openpyxl(ws, row, col + offset)
                                if next_val:
                                    next_str = safe_str_openpyxl(next_val)
                                    for ck in child_keywords:
                                        if ck in next_str:
                                            return col + offset
                        if row + 1 <= ws.max_row:
                            down_val = get_cell_value_openpyxl(ws, row + 1, col)
                            if down_val:
                                down_str = safe_str_openpyxl(down_val)
                                for ck in child_keywords:
                                    if ck in down_str:
                                        return col
                        for offset in range(1, 20):
                            if col + offset <= ws.max_column and row + 1 <= ws.max_row:
                                down_right = get_cell_value_openpyxl(ws, row + 1, col + offset)
                                if down_right:
                                    down_str = safe_str_openpyxl(down_right)
                                    for ck in child_keywords:
                                        if ck in down_str:
                                            return col + offset
    return None

def find_column_direct_openpyxl(ws, keywords, start_row, end_row, max_col=50):
    for row in range(start_row, min(end_row, ws.max_row + 1)):
        for col in range(1, min(max_col, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val:
                val_str = safe_str_openpyxl(val)
                for kw in keywords:
                    if kw in val_str:
                        return col
    return None

def extract_table6_openpyxl(ws, start_row, end_row):
    result = {"cover": None, "plants": []}
    for row in range(start_row, min(end_row, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val and "灌木及高大草本总盖度" in safe_str_openpyxl(val):
                for offset in range(1, 6):
                    right_val = get_cell_value_openpyxl(ws, row, col + offset)
                    num = extract_number_openpyxl(right_val)
                    if num is not None and 0 <= num <= 100:
                        result["cover"] = num
                        break
                if result["cover"] is None:
                    for offset in range(1, 6):
                        down_val = get_cell_value_openpyxl(ws, row + offset, col)
                        num = extract_number_openpyxl(down_val)
                        if num is not None and 0 <= num <= 100:
                            result["cover"] = num
                            break
                break
        if result["cover"] is not None:
            break
    name_row, name_col = None, None
    for row in range(start_row, min(end_row, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            val = get_cell_value_openpyxl(ws, row, col)
            if val and safe_str_openpyxl(val) == "中名":
                name_row, name_col = row, col
                break
        if name_row:
            break
    if name_row:
        height_col = find_parent_child_column_openpyxl(ws, ["高度(cm)", "高度"], ["平均"], start_row, end_row)
        for row in range(name_row + 1, min(end_row, ws.max_row + 1)):
            val = get_cell_value_openpyxl(ws, row, name_col)
            if val:
                name = safe_str_openpyxl(val)
                if len(name) >= 2 and re.search(r'[\u4e00-\u9fff]', name):
                    height = None
                    if height_col:
                        height = extract_number_openpyxl(get_cell_value_openpyxl(ws, row, height_col))
                    result["plants"].append((name, height))
    return result

def extract_table7_plants_openpyxl(ws, start_row, end_row):
    plant_data = {}
    cluster_col = find_column_direct_openpyxl(ws, ["每亩株丛数"], start_row, min(start_row + 15, ws.max_row))
    area_col = find_parent_child_column_openpyxl(ws, ["株丛占面积"], ["每亩面积", "亩面积"], start_row, end_row)
    weight_col = find_parent_child_column_openpyxl(ws, ["平均每亩重", "每亩重"], ["干重"], start_row, end_row)
    if not cluster_col:
        return plant_data
    plant_rows = []
    for row in range(start_row + 3, min(end_row, ws.max_row + 1)):
        val = get_cell_value_openpyxl(ws, row, 2)
        if val:
            name_str = safe_str_openpyxl(val)
            if name_str in ["大", "中", "小", "小计"]:
                continue
            if len(name_str) >= 2 and re.search(r'[\u4e00-\u9fff]', name_str):
                plant_rows.append((name_str, row))
    for i, (plant_name, plant_row) in enumerate(plant_rows):
        next_plant_row = plant_rows[i + 1][1] if i + 1 < len(plant_rows) else end_row
        summary_row = None
        for row in range(plant_row + 1, min(next_plant_row + 10, ws.max_row + 1)):
            for col in range(1, min(ws.max_column + 1, 10)):
                val = get_cell_value_openpyxl(ws, row, col)
                if val and "小计" in safe_str_openpyxl(val):
                    summary_row = row
                    break
            if summary_row:
                break
        if summary_row:
            v1 = extract_number_openpyxl(get_cell_value_openpyxl(ws, summary_row, cluster_col))
            v2 = extract_number_openpyxl(get_cell_value_openpyxl(ws, summary_row, area_col)) if area_col else None
            v3 = extract_number_openpyxl(get_cell_value_openpyxl(ws, summary_row, weight_col)) if weight_col else None
            plant_data[plant_name] = {"cluster": v1, "area": v2, "weight": v3}
    return plant_data

# ==================== 第1步：表7小计求和 ====================
print("\n【第1步】表7小计求和")

def find_subtotal_rows(ws_sheet):
    rows = []
    for row in range(1, min(ws_sheet.max_row, 200)):
        for col in range(1, 5):
            val = ws_sheet.cell(row=row, column=col).value
            if val and isinstance(val, str):
                val_clean = re.sub(r'\s+', '', val)
                if '小计' in val_clean:
                    rows.append(row)
                    break
    return rows

def sum_previous_3_rows(ws_sheet, subtotal_row):
    start_row = max(1, subtotal_row - 3)
    col_sums = {}
    for row in range(start_row, subtotal_row):
        for col in range(1, min(ws_sheet.max_column, 50)):
            val = extract_number(ws_sheet.cell(row=row, column=col).value)
            if val is not None:
                col_sums[col] = col_sums.get(col, 0) + val
    return col_sums

files = list(Path(WORK_DIR).glob("*.xlsx")) + list(Path(WORK_DIR).glob("*.xls"))
files = [f for f in files if not f.name.startswith("~")]

for f in files:
    try:
        wb_temp = load_workbook(f, data_only=True)
        modified = False
        for sheet in ['表7', '表七']:
            if sheet not in wb_temp.sheetnames:
                continue
            ws_temp = wb_temp[sheet]
            for row in find_subtotal_rows(ws_temp):
                sums = sum_previous_3_rows(ws_temp, row)
                for col, total in sums.items():
                    old = ws_temp.cell(row=row, column=col).value
                    if extract_number(old) != total:
                        ws_temp.cell(row=row, column=col).value = total
                        modified = True
        if modified:
            wb_temp.save(f)
        wb_temp.close()
    except Exception as e:
        print(f"  处理 {f.name} 时出错: {e}")

print("[OK] 表7求和完成")

# ==================== 第2步：表67录入 ====================
print("\n【第2步】表67录入（从M列为空的行开始，保护前9行表头）")

wb = load_workbook(TARGET_EXCEL)
ws = wb.active

# 从第 HEADER_ROWS+1 行开始查找（保护表头）
current_row = HEADER_ROWS + 1
max_check_row = min(ws.max_row + 100, 10000) if ws.max_row > 0 else 1000
found_empty = False
while current_row <= max_check_row:
    m_val = ws.cell(row=current_row, column=13).value
    if m_val is None or str(m_val).strip() == '':
        found_empty = True
        break
    current_row += 1

if not found_empty:
    current_row = max(ws.max_row + 1, HEADER_ROWS + 1) if ws.max_row > 0 else HEADER_ROWS + 1

print(f"从第 {current_row} 行开始写入（M列为空）")

next_serial = 1
for row in range(HEADER_ROWS + 1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and str(val).isdigit():
        next_serial = max(next_serial, int(val) + 1)

sample_plants = {}
recorded_rows = []

for src_file in files:
    file_name = src_file.name
    base_name = src_file.stem
    print(f"\n  处理: {file_name}")
    try:
        wb_source = load_workbook(src_file, data_only=True)
        cover = None
        table6_plants = []
        table7_data = {}
        for sheet_name in wb_source.sheetnames:
            if "表6" in sheet_name or "表六" in sheet_name:
                sheet = wb_source[sheet_name]
                s_start = find_first_section_openpyxl(sheet)
                s_end = find_next_section_start_openpyxl(sheet, s_start)
                table6 = extract_table6_openpyxl(sheet, s_start, s_end)
                cover = table6["cover"]
                table6_plants = table6["plants"]
            elif "表7" in sheet_name or "表七" in sheet_name:
                sheet = wb_source[sheet_name]
                s_start = find_first_section_openpyxl(sheet)
                s_end = find_next_section_start_openpyxl(sheet, s_start)
                table7_data = extract_table7_plants_openpyxl(sheet, s_start, s_end)
        wb_source.close()
    except Exception as e:
        print(f"   ❌ 读取失败: {e}")
        continue
    sample_plants[base_name] = [p[0] for p in table6_plants]
    recorded_rows.append(current_row)
    safe_write_cell(ws, current_row, 1, next_serial)
    safe_write_cell(ws, current_row, 2, "样地")
    safe_write_cell(ws, current_row, 12, base_name)
    safe_write_cell(ws, current_row, 13, "样地")
    if cover is not None and 0 <= cover <= 100:
        safe_write_cell(ws, current_row, 29, cover)
    
    y_sum = 0
    z_sum = 0
    aa_sum = 0
    
    for plant_name, height in table6_plants:
        if plant_name in table7_data:
            d = table7_data[plant_name]
            if d.get("cluster"):
                y_sum += d.get("cluster")
            if d.get("area"):
                z_sum += d.get("area")
            if d.get("weight"):
                aa_sum += d.get("weight")
    
    if y_sum > 0:
        safe_write_cell(ws, current_row, 25, round(y_sum, 2) if isinstance(y_sum, float) else y_sum)
    if z_sum > 0:
        safe_write_cell(ws, current_row, 26, round(z_sum, 2) if isinstance(z_sum, float) else z_sum)
    if aa_sum > 0:
        safe_write_cell(ws, current_row, 27, round(aa_sum, 2) if isinstance(aa_sum, float) else aa_sum)
    
    next_serial += 1
    current_row += 1
    for plant_name, height in table6_plants:
        recorded_rows.append(current_row)
        safe_write_cell(ws, current_row, 13, plant_name)
        safe_write_cell(ws, current_row, 14, "灌木")
        if height is not None:
            safe_write_cell(ws, current_row, 28, height)
        if plant_name in table7_data:
            d = table7_data[plant_name]
            if d.get("cluster"):
                safe_write_cell(ws, current_row, 25, d["cluster"])
            if d.get("area"):
                safe_write_cell(ws, current_row, 26, d["area"])
            if d.get("weight"):
                safe_write_cell(ws, current_row, 27, d["weight"])
        current_row += 1
    current_row += 200

wb.save(TARGET_EXCEL)
print("\n[OK] 表67录入完成")

# ==================== 第3步：表345录入 ====================
print("\n【第3步】表345录入（草本数据）")

wb = load_workbook(TARGET_EXCEL)
ws = wb.active

for src_file in files:
    file_name = src_file.stem
    print(f"\n  ========== 处理文件: {file_name} ==========")
    sample_row = None
    for row in range(HEADER_ROWS + 1, ws.max_row + 1):
        if ws.cell(row=row, column=12).value == file_name and ws.cell(row=row, column=13).value == "样地":
            sample_row = row
            break
    if sample_row is None:
        print(f"    未找到样地行，跳过")
        continue
    print(f"    找到样地行: 第{sample_row}行")
    recorded_rows.append(sample_row)
    plant_row = sample_row + 1
    while plant_row < ws.max_row:
        val = ws.cell(row=plant_row, column=13).value
        if val is None or str(val).strip() == '':
            break
        plant_row += 1
    print(f"    草本植物将从第{plant_row}行开始写入")
    
    print(f"\n    --- 处理表5 ---")
    df5, plants5, info5 = process_table5(src_file)
    freq_map = {}
    if df5 is not None and info5 is not None and info5.get("freq_col") is not None:
        freq_map = get_plant_data_mapping(df5, info5["name_col"], info5["name_row"], info5["freq_col"], 200)
        print(f"    频度数据: {len(freq_map)} 条")
    
    print(f"\n    --- 处理表3 ---")
    df3, plants3, info3 = process_table3(src_file)
    if df3 is None:
        print(f"    表3处理失败，跳过该文件")
        continue
    
    for i, plant in enumerate(plants3):
        safe_write_cell(ws, plant_row + i, 13, plant)
        safe_write_cell(ws, plant_row + i, 14, "草本")
        recorded_rows.append(plant_row + i)
    print(f"    写入{len(plants3)}种植物")
    
    if info3["height_cols"]:
        for i, plant in enumerate(plants3):
            data_row = info3["name_row"] + 1 + i
            if data_row >= len(df3):
                break
            for j, h_col in enumerate(info3["height_cols"][:4]):
                if h_col < len(df3.columns):
                    val = df3.iloc[data_row, h_col]
                    if pd.notna(val):
                        safe_write_cell(ws, plant_row + i, 17 + j, val)
    
    if info3["cover_col"] is not None:
        cover_map = get_plant_data_mapping(df3, info3["name_col"], info3["name_row"], info3["cover_col"], 150)
        for i, plant in enumerate(plants3):
            if plant in cover_map:
                safe_write_cell(ws, plant_row + i, 21, cover_map[plant])
    
    abund_sum = 0
    if info3["abund_col"] is not None:
        abund_map = get_plant_data_mapping(df3, info3["name_col"], info3["name_row"], info3["abund_col"], 150)
        for i, plant in enumerate(plants3):
            if plant in abund_map:
                val = abund_map[plant]
                safe_write_cell(ws, plant_row + i, 22, val)
                if isinstance(val, (int, float)):
                    abund_sum += val
    
    for i, plant in enumerate(plants3):
        if plant in freq_map:
            safe_write_cell(ws, plant_row + i, 24, freq_map[plant])
    
    shrub_plants = sample_plants.get(file_name, [])
    for plant in shrub_plants:
        shrub_row = sample_row + 1
        while shrub_row < ws.max_row and ws.cell(row=shrub_row, column=13).value != plant:
            shrub_row += 1
        if shrub_row < ws.max_row and plant in freq_map:
            safe_write_cell(ws, shrub_row, 24, freq_map[plant])
            recorded_rows.append(shrub_row)
    
    if "自然生殖枝高" in info3["special_metrics"]:
        data = info3["special_metrics"]["自然生殖枝高"]
        safe_write_cell(ws, sample_row, data["target_col"], data["value"])
        print(f"    自然生殖枝高: {data['value']}")
    
    if "自然叶层高" in info3["special_metrics"]:
        data = info3["special_metrics"]["自然叶层高"]
        safe_write_cell(ws, sample_row, data["target_col"], data["value"])
        print(f"    自然叶层高: {data['value']}")
    
    if "总盖度" in info3["special_metrics"]:
        data = info3["special_metrics"]["总盖度"]
        safe_write_cell(ws, sample_row, data["target_col"], data["value"])
        print(f"    总盖度: {data['value']}")
    
    print(f"\n    --- 处理表4 ---")
    df4, plants4, info4 = process_table4(src_file)
    all_plants = set(plants3)
    
    if df4 is not None and info4 is not None:
        new_plants4 = [p for p in plants4 if is_valid_plant_name(p) and p not in all_plants]
        for p in new_plants4:
            all_plants.add(p)
        
        if new_plants4:
            write_row = plant_row + len(plants3)
            for i, plant in enumerate(new_plants4):
                safe_write_cell(ws, write_row + i, 13, plant)
                safe_write_cell(ws, write_row + i, 14, "草本")
                recorded_rows.append(write_row + i)
            print(f"    表4新增植物: {len(new_plants4)}种")
        
        weight_map = {}
        if info4.get("weight_col") is not None:
            weight_map = get_plant_data_mapping(df4, info4["name_col"], info4["name_row"], info4["weight_col"], 150)
            print(f"    干重数据: {len(weight_map)}条")
        
        weight_sum = 0
        for i, plant in enumerate(plants3):
            if plant in weight_map:
                val = weight_map[plant]
                if isinstance(val, (int, float)):
                    val = round(val, 2)
                safe_write_cell(ws, plant_row + i, 23, val)
                if isinstance(val, (int, float)):
                    weight_sum += val
        
        if new_plants4:
            for i, plant in enumerate(new_plants4):
                if plant in weight_map:
                    val = weight_map[plant]
                    if isinstance(val, (int, float)):
                        val = round(val, 2)
                    safe_write_cell(ws, write_row + i, 23, val)
                    if isinstance(val, (int, float)):
                        weight_sum += val
                if plant in freq_map:
                    safe_write_cell(ws, write_row + i, 24, freq_map[plant])
        
        safe_write_cell(ws, sample_row, 23, round(weight_sum, 2) if isinstance(weight_sum, (int, float)) else weight_sum)
        print(f"    重量合计: {round(weight_sum, 2) if isinstance(weight_sum, (int, float)) else weight_sum}")
    
    if df5 is not None and info5 is not None:
        all_plants_from_tables = set(plants3)
        if 'new_plants4' in locals():
            all_plants_from_tables.update(new_plants4)
        new_plants5 = [p for p in plants5 if is_valid_plant_name(p) and p not in all_plants_from_tables]
        
        if new_plants5:
            if 'new_plants4' in locals() and new_plants4:
                write_row_5 = plant_row + len(plants3) + len(new_plants4)
            else:
                write_row_5 = plant_row + len(plants3)
            
            for i, plant in enumerate(new_plants5):
                safe_write_cell(ws, write_row_5 + i, 13, plant)
                safe_write_cell(ws, write_row_5 + i, 14, "草本")
                recorded_rows.append(write_row_5 + i)
                if plant in freq_map:
                    safe_write_cell(ws, write_row_5 + i, 24, freq_map[plant])
                    print(f"      写入频度: {plant} -> {freq_map[plant]}")
                else:
                    print(f"      警告: {plant} 没有频度数据")
            print(f"    表5新增植物: {len(new_plants5)}种")
        
        for i, plant in enumerate(plants3):
            if plant in freq_map and ws.cell(row=plant_row + i, column=24).value is None:
                safe_write_cell(ws, plant_row + i, 24, freq_map[plant])
                print(f"      补充频度: {plant} -> {freq_map[plant]}")
    
    safe_write_cell(ws, sample_row, 22, abund_sum)
    print(f"    完成！共{len(all_plants)}种植物")

wb.save(TARGET_EXCEL)
print("\n[OK] 表345录入完成")

# ==================== 第4步：表1录入 ====================
print("\n【第4步】表1录入（新规则：扫描分散单元格，D盟市/E旗县/F乡镇）")

with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write("草普录入数据处理日志\n")
    f.write("=" * 60 + "\n")

write_log("开始处理表1数据...")

target_df = pd.read_excel(TARGET_EXCEL, header=None, dtype=str)
while target_df.shape[1] < 30:
    target_df[target_df.shape[1]] = None

source_files = {}
for f in Path(WORK_DIR).glob("*.xls*"):
    if not f.name.startswith("~"):
        name_without_ext = f.stem.strip()
        source_files[name_without_ext] = f
        source_files[f.name.strip()] = f
write_log(f"找到 {len(set([k for k in source_files.keys() if not k.endswith('.xls')]))} 个源文件")

max_serial = 0
for val in target_df.iloc[:, 0]:
    if pd.notna(val):
        try:
            max_serial = max(max_serial, int(float(str(val))))
        except:
            pass

updated_count = 0
skipped_count = 0
no_file_count = 0

# 从表头之后开始处理
for idx in range(HEADER_ROWS, len(target_df)):
    l_val = target_df.iloc[idx, 11] if 11 < target_df.shape[1] else None
    if pd.isna(l_val) or str(l_val).strip() == '':
        continue
    file_name = str(l_val).strip()
    c_val = target_df.iloc[idx, 2] if 2 < target_df.shape[1] else None
    if pd.notna(c_val) and str(c_val).strip():
        skipped_count += 1
        continue
    if file_name not in source_files:
        write_log(f"行{idx+1}: 找不到源文件 '{file_name}'")
        no_file_count += 1
        continue
    info = extract_table1_info(source_files[file_name])
    a_val = target_df.iloc[idx, 0] if 0 < target_df.shape[1] else None
    if pd.isna(a_val) or str(a_val).strip() == '':
        max_serial += 1
        target_df.iloc[idx, 0] = max_serial
    target_df.iloc[idx, 1] = "样地"
    if info.get('样地号'):
        target_df.iloc[idx, 2] = remove_decimal_zero(info['样地号'])
    if info.get('盟市'):
        target_df.iloc[idx, 3] = info['盟市']
    if info.get('旗县'):
        target_df.iloc[idx, 4] = info['旗县']
    if info.get('乡镇'):
        target_df.iloc[idx, 5] = info['乡镇']
    if info.get('经度') is not None:
        target_df.iloc[idx, 6] = info['经度']
    if info.get('纬度') is not None:
        target_df.iloc[idx, 7] = info['纬度']
    if info.get('海拔'):
        target_df.iloc[idx, 8] = info['海拔']
    if info.get('日期'):
        target_df.iloc[idx, 9] = info['日期']
    if info.get('优势度定名'):
        target_df.iloc[idx, 10] = info['优势度定名']
    write_log(f"行{idx+1}: {file_name} - 盟市(D):{info.get('盟市', '无')}, 旗县(E):{info.get('旗县', '无')}, 乡镇(F):{info.get('乡镇', '无')}")
    updated_count += 1

if updated_count > 0:
    backup = TARGET_EXCEL.replace('.xlsx', f'_备份_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    os.makedirs(os.path.dirname(backup), exist_ok=True)
    target_df.to_excel(TARGET_EXCEL, index=False, header=False, engine='openpyxl')
    write_log(f"✅ 完成！更新 {updated_count} 行，跳过(已有数据) {skipped_count} 行，找不到文件 {no_file_count} 行")
else:
    write_log("没有需要更新的行")

print(f"[OK] 表1录入完成")

# ==================== 第5步：删除空行 ====================
print("\n【第5步】删除本次录入范围内M列为空的行（保护前9行表头）")

wb = load_workbook(TARGET_EXCEL)
ws = wb.active

if recorded_rows:
    min_row = min(recorded_rows)
    max_row = max(recorded_rows)
    print(f"本次录入的行范围: 第 {min_row} 行 到 第 {max_row} 行")
    rows_to_keep = []
    # 保留前 HEADER_ROWS 行表头
    for row in range(1, min(HEADER_ROWS + 1, min_row)):
        rows_to_keep.append(row)
    kept_in_range = 0
    deleted_in_range = 0
    for row in range(min_row, max_row + 1):
        m_val = ws.cell(row=row, column=13).value
        if m_val is not None and str(m_val).strip() != '':
            rows_to_keep.append(row)
            kept_in_range += 1
        else:
            deleted_in_range += 1
    for row in range(max_row + 1, ws.max_row + 1):
        rows_to_keep.append(row)
    all_data = []
    for row in rows_to_keep:
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        all_data.append(row_data)
    ws.delete_rows(1, ws.max_row)
    for new_row_idx, row_data in enumerate(all_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=new_row_idx, column=col_idx, value=value)
    print(f"[OK] 在范围内删除了 {deleted_in_range} 个M列为空的行")
    print(f"[OK] 范围内保留了 {kept_in_range} 个有数据的行")
else:
    print("[提示] 未记录到本次录入的行，将跳过删除空行步骤")

wb.save(TARGET_EXCEL)
print("[OK] 删除空行完成")

print("\n" + "=" * 70)
print("全部完成！")
print(f"表1成功更新: {updated_count} 条")
print(f"日志文件: {LOG_FILE}")
print("=" * 70)